from import_export import resources, fields
from import_export.widgets import ForeignKeyWidget

from .models import Maxsulot, OrderItems, CartItems


class MaxsulotResource(resources.ModelResource):
    class Meta:
        model = Maxsulot
        fields = ('nomi', 'rasm', 'foydalanuvchi__username', 'razmer', 'qoshimcha')


class OrderItemsResource(resources.ModelResource):
    maxsulot = fields.Field(
        column_name='maxsulot',
        attribute='maxsulot',
        widget=ForeignKeyWidget(Maxsulot, 'nomi')
    )

    class Meta:
        model = OrderItems
        fields = ('maxsulot', 'soni', 'foydalanuvchi__username')


class CartItemsResource(resources.ModelResource):
    maxsulot = fields.Field(
        column_name='maxsulot',
        attribute='maxsulot',
        widget=ForeignKeyWidget(Maxsulot, 'nomi')
    )

    class Meta:
        model = CartItems
        fields = ('maxsulot', 'soni', 'foydalanuvchi__username')


# class OrderResource(resources.ModelResource):
#     maxsulotlar = fields.Field(
#         column_name='maxsulotlar',
#         attribute='maxsulotlar',
#         widget=ForeignKeyWidget(OrderItems, 'id')
#     )
#
#     class Meta:
#         model = Order
#         fields = ('maxsulotlar__maxsulot__nomi', 'foydalanuvchi__username', 'jami_maxsulot', 'status', 'bekor_qilish_sababi', 'qoshimcha_rasm', 'qoshimcha_matn', 'created_at', 'updated_at')

# from .models import Order
# from import_export import resources
# from import_export.fields import Field
# from openpyxl import Workbook
# from openpyxl.styles import Alignment, Font
# from openpyxl.utils import get_column_letter
# from io import BytesIO
#
# class OrderResource(resources.ModelResource):
#     maxsulotlar = Field(
#         column_name="Mahsulotlar",
#         attribute="maxsulotlar",
#     )
#
#     class Meta:
#         model = Order
#         fields = (
#             'maxsulotlar',
#             'foydalanuvchi__username',
#             'jami_maxsulot',
#             'status',
#             'qoshimcha_matn',
#             'kimga',
#             'created_at',
#         )
#         export_order = (
#             'maxsulotlar',
#             'foydalanuvchi__username',
#             'jami_maxsulot',
#             'status',
#             'qoshimcha_matn',
#             'kimga',
#             'created_at',
#         )
#
#     def dehydrate_maxsulotlar(self, order):
#         """Format the maxsulotlar field for export."""
#         order_items = order.maxsulotlar.all()
#         return "\n".join([  # Format each item in maxsulotlar
#             f"{item.maxsulot.nomi} ({item.maxsulot.razmer}) x {item.soni}" for item in order_items
#         ])
#
#     def get_export_headers(self):
#         """Rename headers for better readability in the exported file."""
#         headers = super().get_export_headers()
#         headers[headers.index('foydalanuvchi__username')] = "Foydalanuvchi"
#         headers[headers.index('jami_maxsulot')] = "Jami Mahsulot"
#         headers[headers.index('status')] = "Holati"
#         headers[headers.index('qoshimcha_matn')] = "Qo'shimcha Matn"
#         headers[headers.index('kimga')] = "Kimga"
#         headers[headers.index('created_at')] = "Yaratilgan Sana"
#         return headers
#
#     def export_to_xlsx(self, queryset):
#         """Export data to an XLSX file with broader columns and rows."""
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Buyurtmalar"
#
#         # Add headers with bold formatting
#         headers = self.get_export_headers()
#         header_font = Font(bold=True)
#         ws.append(headers)
#
#         # Format headers
#         for col_num, header in enumerate(headers, start=1):
#             cell = ws.cell(row=1, column=col_num)
#             cell.font = header_font
#             cell.alignment = Alignment(horizontal='center', vertical='center')
#
#         # Add rows with data
#         for obj in queryset:
#             row_data = [
#                 "\n".join([  # Format maxsulotlar field
#                     f"{item.maxsulot.nomi} ({item.maxsulot.razmer}) x {item.soni}" for item in obj.maxsulotlar.all()
#                 ]),
#                 obj.foydalanuvchi.username,
#                 obj.jami_maxsulot,
#                 dict(obj._meta.get_field('status').choices).get(obj.status, obj.status),
#                 obj.qoshimcha_matn,
#                 dict(obj._meta.get_field('kimga').choices).get(obj.kimga, obj.kimga),
#                 obj.created_at.strftime("%Y-%m-%d %H:%M:%S") if obj.created_at else "",
#             ]
#             ws.append(row_data)
#
#         # Set a broad, uniform width for all columns
#         broad_width = 30  # Adjust width as needed
#         for col_idx in range(1, len(headers) + 1):
#             ws.column_dimensions[get_column_letter(col_idx)].width = broad_width
#
#         # Set a uniform height for all rows
#         broad_height = 40  # Adjust height as needed
#         for row in ws.iter_rows(min_row=1):  # Include the header row
#             ws.row_dimensions[row[0].row].height = broad_height
#
#             # Ensure text alignment for all cells
#             for cell in row:
#                 cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#
#         # Save the workbook to a BytesIO stream
#         output = BytesIO()
#         wb.save(output)
#         output.seek(0)
#         return output
#

#
# from import_export import resources
# from io import BytesIO
# from openpyxl import load_workbook
# from openpyxl.styles import Alignment
#
#
# class OrderResource(resources.ModelResource):
#     # Existing code...
#
#     def export_to_template_xlsx(self, queryset, template_path):
#         """
#         Export data to an XLSX file using a predefined template.
#         """
#         wb = load_workbook(template_path)
#         ws = wb.active
#
#         # Assuming the template has placeholders like {{ field_name }}
#         for row_num, obj in enumerate(queryset, start=2):  # Start from row 2 (skip headers)
#             ws[f"A{row_num}"] = obj.foydalanuvchi.username
#             ws[f"B{row_num}"] = obj.jami_maxsulot
#             ws[f"C{row_num}"] = dict(obj._meta.get_field('status').choices).get(obj.status, obj.status)
#             ws[f"D{row_num}"] = obj.qoshimcha_matn
#             ws[f"E{row_num}"] = dict(obj._meta.get_field('kimga').choices).get(obj.kimga, obj.kimga)
#             ws[f"F{row_num}"] = obj.created_at.strftime("%Y-%m-%d %H:%M:%S") if obj.created_at else ""
#
#         # Set uniform alignment for all cells
#         for row in ws.iter_rows(min_row=2, max_row=len(queryset) + 1):
#             for cell in row:
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
#
#         # Save the workbook to a BytesIO stream
#         output = BytesIO()
#         wb.save(output)
#         output.seek(0)
#         return output

from import_export import resources
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment


# class OrderResource(resources.ModelResource):
#     def export_to_template_xlsx(self, queryset, template_path):
#         """
#         Export data to an XLSX file using a predefined template.
#         """
#         # Load the template
#         try:
#             wb = load_workbook(template_path)
#         except FileNotFoundError:
#             raise FileNotFoundError(f"Template file not found: {template_path}")
#
#         ws = wb.active
#
#         # Populate the data into the template
#         for row_num, obj in enumerate(queryset, start=2):  # Adjust starting row if needed
#             ws[f"A{row_num}"] = obj.id  # Order ID
#             ws[f"B{row_num}"] = obj.foydalanuvchi.username
#             ws[f"C{row_num}"] = obj.jami_maxsulot
#             ws[f"D{row_num}"] = dict(obj._meta.get_field('status').choices).get(obj.status, obj.status)
#             ws[f"E{row_num}"] = obj.qoshimcha_matn
#             ws[f"F{row_num}"] = dict(obj._meta.get_field('kimga').choices).get(obj.kimga, obj.kimga)
#             ws[f"G{row_num}"] = obj.created_at.strftime("%Y-%m-%d %H:%M:%S") if obj.created_at else ""
#
#         # Set uniform alignment for all cells
#         for row in ws.iter_rows(min_row=2, max_row=len(queryset) + 1):
#             for cell in row:
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
#
#         # Save to a BytesIO stream
#         output = BytesIO()
#         wb.save(output)
#         output.seek(0)
#         return output

#
#
# class OrderResource(resources.ModelResource):
#     def export_to_template_xlsx(self, queryset, template_path):
#         """
#         Export data to an XLSX file using a predefined template.
#         """
#         # Load the template
#         try:
#             wb = load_workbook(template_path)
#         except FileNotFoundError:
#             raise FileNotFoundError(f"Template file not found: {template_path}")
#
#         ws = wb.active
#
#         # Populate the data into the template
#         for row_num, obj in enumerate(queryset, start=2):  # Adjust starting row if needed
#             ws[f"A{row_num}"] = obj.id  # Order ID
#
#             # Format the Products field
#             products = "\n".join([
#                 f"{item.maxsulot.nomi} ({item.maxsulot.razmer}) x {item.soni}"
#                 for item in obj.maxsulotlar.all()
#             ])
#             ws[f"B{row_num}"] = products
#
#             ws[f"C{row_num}"] = obj.foydalanuvchi.username  # User
#             ws[f"D{row_num}"] = obj.jami_maxsulot  # Total Products
#
#             # Format the Status field using choices
#             ws[f"E{row_num}"] = dict(obj._meta.get_field('status').choices).get(obj.status, obj.status)
#
#             ws[f"F{row_num}"] = obj.qoshimcha_matn  # Additional Text
#
#             # Format the To field using choices
#             ws[f"G{row_num}"] = dict(obj._meta.get_field('kimga').choices).get(obj.kimga, obj.kimga)
#
#             # Format the Time field
#             ws[f"H{row_num}"] = obj.created_at.strftime("%Y-%m-%d %H:%M:%S") if obj.created_at else ""
#
#         # Set uniform alignment for all cells
#         for row in ws.iter_rows(min_row=2, max_row=len(queryset) + 1):
#             for cell in row:
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
#
#         # Save to a BytesIO stream
#         output = BytesIO()
#         wb.save(output)
#         output.seek(0)
#         return output
#
# class OrderResource(resources.ModelResource):
#     def export_to_template_xlsx(self, queryset, template_path):
#         """
#         Export data to an XLSX file using a predefined template.
#         All products in an order will be concatenated in one row.
#         """
#         # Load the template
#         try:
#             wb = load_workbook(template_path)
#         except FileNotFoundError:
#             raise FileNotFoundError(f"Template file not found: {template_path}")
#
#         ws = wb.active
#
#         row_num = 2  # Starting row for data
#
#         # Populate the data into the template
#         for obj in queryset:
#             products = obj.maxsulotlar.all()
#
#             # Concatenate product details into a single string
#             product_details = "\n".join([
#                 f"{product.maxsulot.nomi} ({product.maxsulot.razmer}) x {product.soni}"
#                 for product in products
#             ]) if products.exists() else "No Products"
#
#             ws[f"A{row_num}"] = obj.id  # Order ID
#             ws[f"B{row_num}"] = product_details  # Products
#             ws[f"C{row_num}"] = obj.foydalanuvchi.username  # User
#             ws[f"D{row_num}"] = obj.jami_maxsulot  # Total Products
#
#             # Format the Status field using choices
#             ws[f"E{row_num}"] = dict(obj._meta.get_field('status').choices).get(obj.status, obj.status)
#
#             ws[f"F{row_num}"] = obj.qoshimcha_matn  # Additional Text
#
#             # Format the To field using choices
#             ws[f"G{row_num}"] = dict(obj._meta.get_field('kimga').choices).get(obj.kimga, obj.kimga)
#
#             # Format the Time field
#             ws[f"H{row_num}"] = obj.created_at.strftime("%Y-%m-%d %H:%M:%S") if obj.created_at else ""
#
#             row_num += 1
#
#         # Set uniform alignment for all cells
#         for row in ws.iter_rows(min_row=2, max_row=row_num - 1):
#             for cell in row:
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
#
#         # Save to a BytesIO stream
#         output = BytesIO()
#         wb.save(output)
#         output.seek(0)
#         return output





class OrderResource(resources.ModelResource):
    def export_to_template_xlsx(self, queryset, template_path):
        """
        Export data to an XLSX file using a predefined template.
        All products in an order will be concatenated in one row, and the row height will adjust automatically.
        """
        # Load the template
        try:
            wb = load_workbook(template_path)
        except FileNotFoundError:
            raise FileNotFoundError(f"Template file not found: {template_path}")

        ws = wb.active

        row_num = 2  # Starting row for data

        # Populate the data into the template
        for obj in queryset:
            products = obj.maxsulotlar.all()

            # Concatenate product details into a single string
            product_details = "\n".join([
                f"{product.maxsulot.nomi} ({product.maxsulot.razmer}) x {product.soni}"
                for product in products
            ]) if products.exists() else "No Products"

            ws[f"A{row_num}"] = obj.id  # Order ID
            ws[f"B{row_num}"] = product_details  # Products
            ws[f"C{row_num}"] = obj.foydalanuvchi.username  # User
            ws[f"D{row_num}"] = obj.jami_maxsulot  # Total Products

            # Format the Status field using choices
            ws[f"E{row_num}"] = dict(obj._meta.get_field('status').choices).get(obj.status, obj.status)

            ws[f"F{row_num}"] = obj.qoshimcha_matn  # Additional Text

            # Format the To field using choices
            ws[f"G{row_num}"] = dict(obj._meta.get_field('kimga').choices).get(obj.kimga, obj.kimga)

            # Format the Time field
            ws[f"H{row_num}"] = obj.created_at.strftime("%Y-%m-%d %H:%M:%S") if obj.created_at else ""

            # Adjust row height to fit product details
            row_height = max(len(product_details.split("\n")) * 15, 15)  # 15 points per line as a base
            ws.row_dimensions[row_num].height = row_height

            row_num += 1

        # Set uniform alignment for all cells
        for row in ws.iter_rows(min_row=2, max_row=row_num - 1):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Save to a BytesIO stream
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output